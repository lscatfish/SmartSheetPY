"""
这是用于解析xlsx文件的py文件
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Alignment
from globalconstants import GlobalConstants as gc


class XlsxLoad:
    """读取xlsx文件的类"""

    def __init__(self, _path: str):
        self.__path = _path
        self.__sheet = []
        self.__load()

    def __load(self):
        """读取文件"""
        print('xlsx文件读取\"' + self.__path + '\"', end = '')
        wb = load_workbook(self.__path)
        wb = load_workbook(self.path, data_only = True, read_only = True)
        ws = wb.worksheets[0]
        for row in ws.iter_rows(values_only = True):  # 遍历全部
            self.__sheet.append(row)
        wb.close()
        print(' - Done!')

    @property
    def path(self):
        """返回文件路径"""
        return self.__path

    @property
    def sheet(self):
        """返回解析到的sheet"""
        return self.__sheet


class XlsxWrite:
    """xlsx写文件"""

    def __init__(
        self,
        path: str = None,
        sheet: list = None,
        title: str = '',
        widths: list = None,
        height: int = None,
        font_regular: Font = gc.fontRegularGBK,
        font_title: Font = gc.fontTitleGBK,
        font_header: Font = gc.fontHeaderGBK,
        border: Border = gc.borderThinBlack,
        has_border: bool = False,
        has_title: bool = False,
        has_header: bool = False,
        alignment: Alignment = gc.alignmentStd,
    ):
        """
        Parameters
        --------
        path : str
            路径
        sheet : list
            表格
        title : str
            标题的名称
        widths : list
            列宽的组合
        height : int
            行高
        font_regular: Font
            正文字体
        font_title: Font
            标题字体
        font_header: Font
            表头字体
        border: Border
            单元格边框
        has_border: bool
            是否加边框
        has_title: bool
            是否有标题
        has_header: bool
            是否有表头
        alignment: Alignment
            对齐方式
        """
        if widths is None: widths = []
        self.__path = path
        self.__sheet = sheet
        self.__title = title
        self.__fontRegular = font_regular
        self.__fontTitle = font_title
        self.__border = border
        self.__hasBorder = has_border
        self.__hasTitle = has_title
        self.__alignment = alignment
        self.__hasTitle = has_title
        self.__hasHeader = has_header
        self.__fontHeader = font_header
        self.__widths = widths
        self.__height = height

    def can_write(self) -> bool:
        """检查能否开始写入表格"""
        if self.__path is None or self.__path == '': return False
        if self.__sheet is None or self.__sheet == []: return False
        if self.__title == '' and self.__hasTitle: return False
        return True

    def write(self, ifp: bool = False) -> bool:
        """写入文件"""
        width_default = 8.43
        if not self.can_write(): return False
        if ifp: print('write xlsx file as \"' + self.__path + '\"', end = '')
        # 创建wb
        wb = Workbook()
        ws = wb.active
        if self.__hasTitle: ws.title = self.__title
        for row in self.__sheet:
            ws.append(row)
        for i in range(1, ws.max_row + 1):
            ws.row_dimensions[i].height = self.__height
        for i in range(1, ws.max_column + 1):
            col_letter = ws.cell(row = 1, column = i).column_letter
            if i < len(self.__widths) and self.__widths[i - 1] > 0:
                ws.column_dimensions[col_letter].width = self.__widths[i - 1]
                width_default = self.__widths[i - 1]
            elif i == len(self.__widths):
                if self.__widths[i - 1] <= 0:
                    ws.column_dimensions[col_letter].width = width_default
                else:
                    ws.column_dimensions[col_letter].width = self.__widths[i - 1]
                    width_default = self.__widths[i - 1]
            else:
                ws.column_dimensions[col_letter].width = width_default

        for row in ws.iter_rows(values_only = False):
            for cell in row:
                cell.alignment = self.__alignment
                if self.__hasBorder: cell.border = self.__border
                cell.font = self.__fontRegular
        if self.__hasHeader:
            for cell in ws[1]:
                cell.font = self.__fontHeader
        if self.__hasTitle:
            ws.insert_rows(1)
            ws.merge_cells(start_row = 1, end_row = 1, start_column = 1, end_column = ws.max_column)
            ws.cell(row = 1, column = 1).value = self.__title
            ws.cell(row = 1, column = 1).font = self.__fontTitle
            ws.cell(row = 1, column = 1).alignment=self.__alignment
        wb.save(self.__path)
        wb.close()
        if ifp: print(' - Done!')
        return True

    @property
    def path(self):
        return self.__path

    @path.setter
    def path(self, value: str):
        if isinstance(value, str):
            self.__path = value
        else:
            raise TypeError("路径必须是str类型")

    @property
    def sheet(self):
        return self.__sheet

    @sheet.setter
    def sheet(self, value: list):
        if isinstance(value, list):
            self.__sheet = value
        else:
            raise TypeError("sheet必须是list类型")

    @property
    def hasTitle(self):
        return self.__hasTitle

    @hasTitle.setter
    def hasTitle(self, i_hasTitle: bool):
        if isinstance(i_hasTitle, bool):
            self.__hasTitle = i_hasTitle
        else:
            raise ValueError("hasTitle必须是bool类型")

    @property
    def title(self):
        return self.__title

    @title.setter
    def title(self, i_title: str):
        if isinstance(i_title, str):
            self.__title = i_title
            self.__hasTitle = True
        else:
            raise ValueError("标题必须是字符串类型")

    @property
    def fontRegular(self):
        return self.__fontRegular

    @fontRegular.setter
    def fontRegular(self, fontRegular: Font):
        self.__fontRegular = fontRegular

    @property
    def fontTitle(self):
        return self.__fontTitle

    @fontTitle.setter
    def fontTitle(self, fontTitle: Font):
        self.__fontTitle = fontTitle
        self.__hasTitle = True

    @property
    def border(self):
        return self.__border

    @border.setter
    def border(self, border: Border):
        self.__border = border
        self.__hasBorder = True

    @property
    def hasBorder(self):
        return self.__hasBorder

    @hasBorder.setter
    def hasBorder(self, hasBorder: bool):
        self.__hasBorder = hasBorder

    @property
    def alignment(self):
        return self.__alignment

    @alignment.setter
    def alignment(self, alignment: Alignment):
        self.__alignment = alignment

    @property
    def widths(self):
        return self.__widths

    @widths.setter
    def widths(self, widths: list):
        self.__widths = widths

    @property
    def height(self):
        return self.__height

    @height.setter
    def height(self, height: int):
        self.__height = height

    @property
    def hasHeader(self):
        return self.__hasHeader

    @hasHeader.setter
    def hasHeader(self, hasHeader: bool):
        self.__hasHeader = hasHeader

    @property
    def fontHeader(self):
        return self.__fontHeader

    @fontHeader.setter
    def fontHeader(self, fontHeader: Font):
        self.__fontHeader = fontHeader
        self.__hasHeader = True



# content = [('we', 'fhds', 'fjjf'),
#            ('df', 'dfdsafds', 'dfasd', 'dfsdafsda', 'dsfsad','房键'),
#            ('dsafdasf', 'dfsda', 'dfrgioehbn'),
#            ['dfadjonf', 'dfnfd', 'dsfnsnj', 'dsnfiu']]
# a = XlsxWrite(path = 'text.xlsx', sheet = content, widths = [35, 25])
# a.title='myTest.xlsx'
# if a.can_write(): a.write(True)
