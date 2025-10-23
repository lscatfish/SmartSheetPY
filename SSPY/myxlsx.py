# -*- coding: utf-8 -*-
"""
这是用于解析xlsx文件的py文件
如果你单独使用此脚本，我们推荐你劫持ppocr的下载器以及相关子线程
If you use this script alone, we recommend that you hijack the ppocr downloader and the related subthreads.
"""
import copy
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Alignment
from .globalconstants import GlobalConstants as gc
from .PersonneInformation import DefPerson


def trans_list_to_person(
    header: tuple | list,
    in_info: list,
    classname: str = None,
    inkey_as_sub: bool = False,
    stdkey_as_sub: bool = True) -> DefPerson | None:
    per = DefPerson()
    info = copy.deepcopy(in_info)
    if classname is not None:
        per.classname = classname
    for i in range(len(header)):
        if i >= len(info): break
        if header[i] == '序号': continue
        per.set_information(
            header[i],
            str(info[i]).strip(),
            inkey_as_sub = inkey_as_sub,
            stdkey_as_sub = stdkey_as_sub)
    if per.classname is None or per.classname == "" or per.classname == 'None': return None
    if per.name == '' and per.studentID == '': return None
    per.optimize()
    return per


def get_header_from_xlsx(in_sheet: list, stdkey_as_sub: bool = True) -> tuple | None:
    """
    从in_sheet中获取表头，并返回去除表头后的列表和表头元组

    Parameters:
        stdkey_as_sub: 自字典库判断表头
        in_sheet: 包含表格数据的列表（每行是一个子列表）

    Returns:
        元组 (header, new_sheet)，其中：
        - header: 表头元组（若找到），否则为 None
        - new_sheet: 去除表头后的列表（若找到表头），否则为原列表
    """
    header = None
    has_found = False  # 是否找到表头
    new_sheet = []
    """表头重复打印也会被清除"""
    for i, row in enumerate(in_sheet):
        if not has_found:
            # 检查当前行是否包含表头标识（通过DefPerson.get_stdkey判断）
            for cell in row:
                if DefPerson.get_stdkey(cell, stdkey_as_sub = stdkey_as_sub) is not None:
                    header = tuple(row)  # 表头所在行转换为元组
                    has_found = True
                    break  # 找到表头后跳出单元格循环
            # 如果找到表头，不将当前行加入新列表
            if has_found:
                continue
        # 非表头行加入新列表
        new_sheet.append(row)

    return header, new_sheet


def clear_empty_lines(in_sheet: list[list]) -> list:
    """清理空行"""
    if len(in_sheet) == 0:
        return []
    new_sheet = []
    for row in in_sheet:
        for cell in row:
            if not (cell == '' or cell is None or str(cell).strip() == '' or str(cell) == 'None'):
                new_sheet.append(row)
                break
    return new_sheet


class XlsxLoad:
    """读取xlsx文件的类"""

    def __init__(
        self,
        _path: str,
        const_classname: bool = True,
        classname: str = None,
        ifp: bool = False,
        header: list[str] = None) -> None:
        from .myfolder import split_filename_and_extension
        self.__path = _path
        self.__sheets = []
        self.__ifp = ifp
        self.__const_classname = const_classname
        if const_classname:
            if classname is None:
                self.__classname = split_filename_and_extension(self.__path)[0]
            else:
                self.__classname = classname
        self.__header = header
        self.__load()

    def __load(self):
        """读取文件"""
        if self.__ifp:
            print('xlsx文件读取\"' + self.__path + '\"', end = '')
        try:
            wb = load_workbook(self.__path, data_only = True, read_only = True)
        except Exception as e :
            print(f'xlsx文件"{self.__path}"解析出错：{e} ，已跳过...')
            return

        for ws in wb.worksheets:
            sh = []
            for row in ws.iter_rows(values_only = True):  # 遍历全部
                sh.append(row)
            self.__sheets.append(sh)
        wb.close()
        if self.__ifp:
            print(' - Done!')

    @property
    def path(self):
        """返回文件路径"""
        return self.__path

    @property
    def sheets(self):
        """返回解析到的sheet"""
        return copy.deepcopy(self.__sheets)

    def get_personList(
        self,
        inkey_as_sub: bool = False,
        stdkey_as_sub: bool = False):
        pers: list[DefPerson] = []
        header, only_sheet = get_header_from_xlsx(self.sheets[0], stdkey_as_sub = stdkey_as_sub)
        if self.__header is None and header is None:
            print('文件 \"' + self.__path + '\" 未找到表头')
            return pers
        elif self.__header is not None and header is None:
            header = self.__header
        if self.__const_classname:
            for row in only_sheet:
                p = trans_list_to_person(
                    header, row,
                    classname = self.__classname,
                    inkey_as_sub = inkey_as_sub,
                    stdkey_as_sub = stdkey_as_sub)
                if p is not None:
                    pers.append(p)
        else:
            for row in only_sheet:
                p = trans_list_to_person(
                    header, row,
                    inkey_as_sub = inkey_as_sub,
                    stdkey_as_sub = stdkey_as_sub)
                if p is not None:
                    pers.append(p)
        return pers


class XlsxWrite:
    """xlsx写文件"""

    def __init__(
        self,
        path: str = None,
        sheet: list = None,
        widths: list[float] = None,
        height: float = None,
        title: str = '',
        height_title: float = 40,
        font_regular: Font = gc.fontRegularGBK,
        font_title: Font = gc.fontTitleGBK,
        font_header: Font = gc.fontHeaderGBK,
        border: Border = gc.borderThinBlack,
        has_border: bool = False,
        has_title: bool = False,
        has_header: bool = False,
        alignment: Alignment = gc.alignmentStd,
    ):
        """
        Parameters
        --------
        path : str
            路径
        sheet : list
            表格
        title : str
            标题的名称
        widths : list
            列宽的组合
        height : int
            行高
        font_regular: Font
            正文字体
        font_title: Font
            标题字体
        font_header: Font
            表头字体
        border: Border
            单元格边框
        has_border: bool
            是否加边框
        has_title: bool
            是否有标题
        has_header: bool
            是否有表头
        alignment: Alignment
            对齐方式
        """
        if widths is None: widths = []
        self.__path = path
        self.__sheet = copy.deepcopy(sheet)
        self.__title = title
        self.__fontRegular = font_regular
        self.__fontTitle = font_title
        self.__border = border
        self.__hasBorder = has_border
        self.__hasTitle = has_title
        self.__alignment = alignment
        self.__hasTitle = has_title
        self.__hasHeader = has_header
        self.__fontHeader = font_header
        self.__widths = copy.deepcopy(widths)
        self.__height = height
        self.__heightTitle = height_title

    @property
    def can_write(self) -> bool:
        """检查能否开始写入表格"""
        if self.__path is None or self.__path == '': return False
        if self.__sheet is None or self.__sheet == []: return False
        if self.__title == '' and self.__hasTitle: return False
        return True

    def write(self, ifp: bool = False) -> bool:
        """写入文件"""
        width_default = 8.43
        if not self.can_write: return False
        if ifp: print('write xlsx file as \"' + self.__path + '\"', end = '')
        # 创建wb
        wb = Workbook()
        ws = wb.active
        if self.__hasTitle: ws.title = self.__title
        for row in self.__sheet:
            ws.append(row)
        if self.__height is not None:
            for i in range(1, ws.max_row + 2):  # 这里我不晓得为什么是2
                ws.row_dimensions[i].height = self.__height

        if self.__widths is not None and len(self.__widths) > 0:
            for i in range(1, ws.max_column + 1):
                col_letter = ws.cell(row = 1, column = i).column_letter
                if i < len(self.__widths) and self.__widths[i - 1] > 0:
                    ws.column_dimensions[col_letter].width = self.__widths[i - 1]
                    width_default = self.__widths[i - 1]
                elif i == len(self.__widths):
                    if self.__widths[i - 1] <= 0:
                        ws.column_dimensions[col_letter].width = width_default
                    else:
                        ws.column_dimensions[col_letter].width = self.__widths[i - 1]
                        width_default = self.__widths[i - 1]
                else:
                    ws.column_dimensions[col_letter].width = width_default

        for row in ws.iter_rows(values_only = False):
            for cell in row:
                cell.alignment = self.__alignment
                if self.__hasBorder: cell.border = self.__border
                cell.font = self.__fontRegular
        if self.__hasHeader:
            for cell in ws[1]:
                cell.font = self.__fontHeader
        if self.__hasTitle:
            ws.insert_rows(1)
            ws.merge_cells(start_row = 1, end_row = 1, start_column = 1, end_column = ws.max_column)
            ws.cell(row = 1, column = 1).value = self.__title
            ws.cell(row = 1, column = 1).font = self.__fontTitle
            ws.cell(row = 1, column = 1).alignment = self.__alignment
            ws.row_dimensions[1].height = self.__heightTitle if self.__heightTitle > 0 else 25
        wb.save(self.__path)
        wb.close()
        if ifp: print(' - Done!')
        return True

    @property
    def path(self):
        return self.__path

    @path.setter
    def path(self, value: str):
        if isinstance(value, str):
            self.__path = value
        else:
            raise TypeError("路径必须是str类型")

    @property
    def sheet(self):
        return copy.deepcopy(self.__sheet)

    @sheet.setter
    def sheet(self, value: list):
        if isinstance(value, list):
            self.__sheet = copy.deepcopy(value)
        else:
            raise TypeError("sheet必须是list类型")

    @property
    def hasTitle(self):
        return self.__hasTitle

    @hasTitle.setter
    def hasTitle(self, i_hasTitle: bool):
        if isinstance(i_hasTitle, bool):
            self.__hasTitle = i_hasTitle
        else:
            raise ValueError("hasTitle必须是bool类型")

    @property
    def title(self):
        return self.__title

    @title.setter
    def title(self, i_title: str):
        if isinstance(i_title, str):
            self.__title = i_title
            self.__hasTitle = True
        else:
            raise ValueError("标题必须是字符串类型")

    @property
    def fontRegular(self):
        return copy.deepcopy(self.__fontRegular)

    @fontRegular.setter
    def fontRegular(self, fontRegular: Font):
        self.__fontRegular = copy.deepcopy(fontRegular)

    @property
    def fontTitle(self):
        return copy.deepcopy(self.__fontTitle)

    @fontTitle.setter
    def fontTitle(self, fontTitle: Font):
        self.__fontTitle = copy.deepcopy(fontTitle)
        self.__hasTitle = True

    @property
    def border(self):
        return copy.deepcopy(self.__border)

    @border.setter
    def border(self, border: Border):
        self.__border = copy.deepcopy(border)
        self.__hasBorder = True

    @property
    def hasBorder(self):
        return self.__hasBorder

    @hasBorder.setter
    def hasBorder(self, hasBorder: bool):
        self.__hasBorder = hasBorder

    @property
    def alignment(self):
        return copy.deepcopy(self.__alignment)

    @alignment.setter
    def alignment(self, alignment: Alignment):
        self.__alignment = copy.deepcopy(alignment)

    @property
    def widths(self):
        return copy.deepcopy(self.__widths)

    @widths.setter
    def widths(self, widths: list):
        self.__widths = copy.deepcopy(widths)

    @property
    def height(self):
        return self.__height

    @height.setter
    def height(self, height: float):
        if height > 0:
            self.__height = height
        else:
            self.__height = 25

    @property
    def heightTitle(self):
        return self.__heightTitle

    @heightTitle.setter
    def heightTitle(self, value: float):
        if value <= 0:
            self.__heightTitle = 25
            self.__hasTitle = False
        else:
            self.__heightTitle = value
            self.__hasTitle = True

    @property
    def hasHeader(self):
        return self.__hasHeader

    @hasHeader.setter
    def hasHeader(self, hasHeader: bool):
        self.__hasHeader = hasHeader

    @property
    def fontHeader(self):
        return copy.deepcopy(self.__fontHeader)

    @fontHeader.setter
    def fontHeader(self, fontHeader: Font):
        self.__fontHeader = copy.deepcopy(fontHeader)
        self.__hasHeader = True
